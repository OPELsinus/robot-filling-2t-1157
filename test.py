import datetime
import os
import random
import shutil
import sys
import time
import uuid
from pathlib import Path
from time import sleep

import numpy as np
import openpyxl
import pandas as pd

import psycopg2 as psycopg2
from openpyxl import load_workbook
from pywinauto import keyboard

from config import logger, robot_name, db_host, db_port, db_name, db_user, db_pass, owa_username, owa_password, working_path, download_path, tg_token, chat_id
from rpamini import Web, App
from tools import update_credentials, send_message_to_tg, send_file_to_tg
from pyautogui import screenshot

from openpyxl import load_workbook

from openpyxl.styles import PatternFill, Alignment

dick1 = {'РИС ВЕС': 'Рис',
         'КРУПА ГРЕЧНЕВАЯ ВЕС': 'гречневая',
         'МАСЛО ПОДСОЛНЕЧНОЕ': 'подсолнечн',
         'КАПУСТА': 'белокоч',
         'ЛУК, ЧЕСНОК': 'репчатый',
         'МОРКОВЬ': 'Морковь',
         'КАРТОФЕЛЬ': 'Картофель',
         'САХАР ПЕСОК': 'Сахар',
         'СОЛЬ ОБЫЧНАЯ': 'Соль'}

groups = ['Объем розничной торговли',
          'Товарные запасы на конец отчетного месяца',
          'Рис',
          'гречневая',
          'подсолнечн',
          'белокоч',
          'репчатый',
          'Морковь',
          'Картофель',
          'Сахар',
          'Соль']


def drop_table():
    conn = psycopg2.connect(host=db_host, port=db_port, database=db_name, user=db_user, password=db_pass)
    table_create_query = f'''
            drop table robot.{robot_name.replace("-", "_")}
            '''
    c = conn.cursor()
    c.execute(table_create_query)

    conn.commit()
    c.close()
    conn.close()


def sql_create_table():
    conn = psycopg2.connect(host=db_host, port=db_port, database=db_name, user=db_user, password=db_pass)
    table_create_query = f'''
        CREATE TABLE IF NOT EXISTS ROBOT.{robot_name.replace("-", "_")}_1 (
            started_time timestamp,
            ended_time timestamp,
            store_id int PRIMARY KEY,
            store_name text,
            status text,
            error_reason text,
            error_saved_path text,
            execution_time text,
            fact1 text,
            fact2 text,
            fact3 text,
            site1 text,
            site2 text,
            site3 text
            )
        '''
    c = conn.cursor()
    c.execute(table_create_query)

    conn.commit()
    c.close()
    conn.close()


def delete_by_id(id):
    conn = psycopg2.connect(host=db_host, port=db_port, database=db_name, user=db_user, password=db_pass)
    table_create_query = f'''
                DELETE FROM ROBOT.{robot_name.replace("-", "_")}_1 WHERE id = '{id}'
                '''
    c = conn.cursor()
    c.execute(table_create_query)
    conn.commit()
    c.close()
    conn.close()


def get_all_data():
    conn = psycopg2.connect(host=db_host, port=db_port, database=db_name, user=db_user, password=db_pass)
    table_create_query = f'''
            SELECT * FROM ROBOT.{robot_name.replace("-", "_")}_1
            order by started_time desc
            '''
    cur = conn.cursor()
    cur.execute(table_create_query)

    df1 = pd.DataFrame(cur.fetchall())
    df1.columns = ['started_time', 'ended_time', 'store_id', 'name', 'status', 'error_reason', 'error_saved_path', 'execution_time', 'fact1', 'fact2', 'fact3', 'site1', 'site2', 'site3']

    cur.close()
    conn.close()

    return df1


def get_data_by_name(store_name):
    conn = psycopg2.connect(host=db_host, port=db_port, database=db_name, user=db_user, password=db_pass)
    table_create_query = f'''
            SELECT * FROM ROBOT.{robot_name.replace("-", "_")}_1
            where store_name = '{store_name}'
            order by started_time desc
            '''
    cur = conn.cursor()
    cur.execute(table_create_query)

    df1 = pd.DataFrame(cur.fetchall())
    # df1.columns = ['started_time', 'ended_time', 'store_id', 'name', 'status', 'error_reason', 'error_saved_path', 'execution_time']

    cur.close()
    conn.close()

    return len(df1)


def insert_data_in_db(started_time, store_id, store_name, status, error_reason, error_saved_path, execution_time, fact1, fact2, fact3, site1, site2, site3):

    conn = psycopg2.connect(host=db_host, port=db_port, database=db_name, user=db_user, password=db_pass)

    print('Started inserting')
    query_delete_id = f"""
            delete from ROBOT.{robot_name.replace("-", "_")}_1 where store_id = '{store_id}'
        """
    query_delete = f"""
        delete from ROBOT.{robot_name.replace("-", "_")}_1 where store_name = '{store_name}'
    """
    query = f"""
        INSERT INTO ROBOT.{robot_name.replace("-", "_")}_1 (started_time, ended_time, store_id, store_name, status, error_reason, error_saved_path, execution_time, fact1, fact2, fact3, site1, site2, site3)
        VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
    """

    values = (
        started_time,
        datetime.datetime.now().strftime("%d.%m.%Y %H:%M:%S.%f"),
        store_id,
        store_name,
        status,
        error_reason,
        error_saved_path,
        execution_time,
        fact1,
        fact2,
        fact3,
        site1,
        site2,
        site3
    )
    print(values)

    cursor = conn.cursor()

    cursor.execute(query_delete)
    # conn.autocommit = True
    try:
        cursor.execute(query_delete)
        cursor.execute(query_delete_id)
    except Exception as e:
        print('GOVNO', e)
        pass
    try:
        cursor.execute(query, values)
    except Exception as e:
        conn.rollback()
        print(f"Error: {e}")

    conn.commit()

    cursor.close()
    conn.close()


def get_all_branches_with_codes():
    import psycopg2
    import pandas as pd

    conn = psycopg2.connect(dbname='adb', host='172.16.10.22', port='5432',
                            user='rpa_robot', password='Qaz123123+')

    cur = conn.cursor(name='1583_first_part')

    query = f"""
        select db.id_sale_object, ds.source_store_id, ds.store_name, ds.sale_obj_name 
        from dwh_data.dim_branches db
        left join dwh_data.dim_store ds on db.id_sale_object = ds.sale_source_obj_id
        where ds.store_name like '%Торговый%' and current_date between ds.datestart and ds.dateend
        group by db.id_sale_object, ds.source_store_id, ds.store_name, ds.sale_obj_name
        order by ds.source_store_id
    """

    cur.execute(query)

    print('Executed')

    df1 = pd.DataFrame(cur.fetchall())
    df1.columns = ['branch_id', 'store_id', 'store_name', 'store_normal_name']

    cur.close()
    conn.close()

    return df1


def sign_ecp(ecp):
    logger.info('Started ECP')
    # print('Kek')
    logger.info(f'KEY: {ecp}')
    # print(f'KEY: {ecp}')
    app = App('')

    el = {"title": "Открыть файл", "class_name": "SunAwtDialog", "control_type": "Window",
          "visible_only": True, "enabled_only": True, "found_index": 0, "parent": None}

    if app.wait_element(el, timeout=30):

        keyboard.send_keys(ecp.replace('(', '{(}').replace(')', '{)}'), pause=0.01, with_spaces=True)
        sleep(0.05)
        keyboard.send_keys('{ENTER}')

        if app.wait_element({"title_re": "Формирование ЭЦП.*", "class_name": "SunAwtDialog", "control_type": "Window",
                             "visible_only": True, "enabled_only": True, "found_index": 0, "parent": None}, timeout=30):
            app.find_element({"title_re": "Формирование ЭЦП.*", "class_name": "SunAwtDialog", "control_type": "Window",
                              "visible_only": True, "enabled_only": True, "found_index": 0, "parent": None}).type_keys('Aa123456')

            # keyboard.send_keys('Aa123456')
            sleep(2)

            keyboard.send_keys('{ENTER}')
            sleep(3)

            keyboard.send_keys('{ENTER}')
            app = None
            logger.info('Finished ECP')
        else:
            logger.info('Quit mazafaka1')
            app = None
            return 'broke'
    else:
        logger.info('Quit mazafaka')
        # try:
        #     app.find_element({"title": "Закрыть", "class_name": "", "control_type": "Button",
        #                       "visible_only": True, "enabled_only": True, "found_index": 0}, ti).click()
        #     app.quit()
        # except:
        #     pass
        app = None
        return 'broke'
    # app.quit()


def save_screenshot(store):
    scr = screenshot()
    scr_path = str(os.path.join(download_path, str(store + '_HUETA.png')))
    scr.save(scr_path)

    return scr_path


def wait_loading(web, xpath):
    print('Started loading')
    ind = 0
    element = ''
    while True:
        # if ind == 0:
        #     web.find_element(xpath).click()
        try:
            print(web.lolus('//*[@id="loadmask-1315"]'))
            if web.lolus('//*[@id="loadmask-1315"]') == '':
                element = ''
            if (element == '' and web.lolus('//*[@id="loadmask-1315"]') == 'none') or (ind >= 1000 and element == ''):
                print('Loaded')
                sleep(0.5)
                break
            ind += 1
        except:
            pass


def create_and_send_final_report():

    if __name__ == '__main__':

        df = get_all_data()

        df[df.columns[-1]] = df[df.columns[-1]].astype(float)
        df[df.columns[-1]] = df[df.columns[-1]].round()

        df.columns = ['Время начала', 'Время окончания', 'Номер филиала', 'Название филиала', 'Статус', 'Причина ошибки', 'Пусть сохранения скриншота', 'Время исполнения (сек)']

        df.to_excel('result.xlsx', index=False)

        workbook = load_workbook('result.xlsx')
        sheet = workbook.active

        red_fill = PatternFill(start_color="FFA864", end_color="FFA864", fill_type="solid")
        green_fill = PatternFill(start_color="A6FF64", end_color="A6FF64", fill_type="solid")

        for cell in sheet['E']:
            if cell.value == 'failed':
                cell.fill = red_fill
            if cell.value == 'success':
                cell.fill = green_fill

        for col in 'ABCDGH':

            max_length = max(len(str(cell.value)) for cell in sheet[col])

            if col == 'A' or col == 'B':
                max_length -= 3
            if col == 'D':
                max_length += 5
            if col == 'A':
                max_length -= 3

            sheet.column_dimensions[col].width = max_length

        for col in 'ABCDGEFGH':
            for cell in sheet[col]:
                cell.alignment = Alignment(horizontal='center')

        workbook.save('result.xlsx')

        send_file_to_tg(tg_token, chat_id, 'Дросте', 'result.xlsx')


def wait_image_loaded():
    found = False
    while True:
        for file in os.listdir(download_path):
            if '.jpg' in file and 'crdownload' not in file:
                shutil.move(os.path.join(download_path, file), os.path.join(os.path.join(download_path, 'Отчёты 2т'), branch + '_1.jpg'))
                print(file)
                found = True
                break
        if found:
            break


def save_and_send(web, save):
    if save:
        web.execute_script_click_xpath("//span[text() = 'Сохранить']")
        sleep(1)
        print('Clicked Save')
        if web.wait_element("//span[text() = 'Сохранить отчет и Удалить другие']", timeout=5):
            web.execute_script_click_xpath("//span[text() = 'Сохранить отчет и Удалить другие']")
    print('Clicking Send')
    web.execute_script_click_xpath("//span[text() = 'Отправить']")
    print('Clicked Send')
    web.wait_element("//input[@value = 'Персональный компьютер']", timeout=30)
    web.execute_script_click_xpath("//input[@value = 'Персональный компьютер']")


def start_single_branch(filepath, store, values_first_part, values_second_part, store_id):
    print('Started web')
    web = Web()
    web.run()
    web.get('https://cabinet.stat.gov.kz/')
    logger.info('Check-1')
    web.wait_element('//*[@id="idLogin"]')
    web.find_element('//*[@id="idLogin"]').click()

    web.wait_element('//*[@id="button-1077-btnEl"]')
    web.find_element('//*[@id="button-1077-btnEl"]').click()

    web.wait_element('//*[@id="lawAlertCheck"]')
    web.find_element('//*[@id="lawAlertCheck"]').click()

    sleep(0.5)
    web.find_element('//*[@id="loginButton"]').click()

    logger.info('Check-2')
    ecp_auth = ''
    ecp_sign = ''
    for files in os.listdir(filepath):
        if 'AUTH' in files:
            ecp_auth = os.path.join(filepath, files)
        if 'GOST' in files:
            ecp_sign = os.path.join(filepath, files)

    sleep(1)
    sign_ecp(ecp_auth)

    logged_in = web.wait_element('//*[@id="idLogout"]/a', timeout=30)
    # sleep(1000)
    if logged_in:
        if web.find_element("//a[text() = 'Выйти']"):

            if web.wait_element("//span[contains(text(), 'Пройти позже')]", timeout=5):
                try:
                    web.find_element("//span[contains(text(), 'Пройти позже')]").click()
                except:
                    save_screenshot(store)
            logger.info('Check0')
            if web.wait_element('//*[@id="dontAgreeId-inputEl"]', timeout=5):
                web.find_element('//*[@id="dontAgreeId-inputEl"]').click()
                sleep(0.3)
                web.find_element('//*[@id="saveId-btnIconEl"]').click()
                sleep(1)
                web.find_element('//*[@id="ext-gen1893"]').click()
                web.find_element('//*[@id="boundlist-1327-listEl"]/ul/li').click()

                sleep(1)
                web.find_element('//*[@id="button-1326-btnIconEl"]').click()
                print('Done lol')
                sign_ecp(ecp_sign)

                try:
                    web.wait_element("//span[contains(text(), 'Пройти позже')]", timeout=5)
                    web.find_element("//span[contains(text(), 'Пройти позже')]").click()

                except:
                    pass
            logger.info('Check1')
            web.wait_element('//*[@id="tab-1168-btnInnerEl"]')
            web.find_element('//*[@id="tab-1168-btnInnerEl"]').click()

            # sleep(0.7)

            web.wait_element('//*[@id="radio-1131-boxLabelEl"]')

            if web.wait_element("//span[contains(text(), 'Пройти позже')]", timeout=1.5):
                web.execute_script_click_xpath("//span[contains(text(), 'Пройти позже')]")
            sleep(1)

            # ? Check if 2T exists
            for i in range(3):

                wait_loading(web, '//*[@id="loadmask-1315"]')
                if web.wait_element("//span[contains(text(), 'Пройти позже')]", timeout=1.5):
                    web.execute_script_click_xpath("//span[contains(text(), 'Пройти позже')]")
                if web.wait_element("//div[contains(text(), '2-торговля')]", timeout=3):
                    web.find_element("//div[contains(text(), '2-торговля')]").click()
                else:
                    if i < 2:

                        if web.wait_element("//span[contains(text(), 'Пройти позже')]", timeout=.5):
                            web.execute_script_click_xpath("//span[contains(text(), 'Пройти позже')]")
                        web.find_element('//*[@id="radio-1133-boxLabelEl"]').click()
                        wait_loading(web, '//*[@id="loadmask-1315"]')
                        web.refresh()

                    else:
                        saved_path = save_screenshot(store)
                        web.close()
                        web.quit()

                        print('Return those shit')
                        return ['Нет 2-т', saved_path]
            logger.info('Check2')
            sleep(0.5)

            web.find_element('//*[@id="createReportId-btnIconEl"]').click()

            sleep(1)

            # ? Switch to the second window
            web.driver.switch_to.window(web.driver.window_handles[-1])

            web.wait_element('//*[@id="td_select_period_level_1"]/span')
            web.execute_script_click_js("#btn-opendata")
            sleep(0.3)

            if web.lolus('/html/body/div[7]') == 'block':

                web.find_element('/html/body/div[7]/div[11]/div/button[2]').click()

                saved_path = save_screenshot(store)
                web.close()
                web.quit()

                print('Return that shit')
                return ['Выскочила хуета', saved_path]
            logger.info('Check3')
            web.wait_element('//*[@id="sel_statcode_accord"]/div/p/b[1]')
            web.execute_script_click_js("body > div:nth-child(16) > div.ui-dialog-buttonpane.ui-widget-content.ui-helper-clearfix > div > button:nth-child(1) > span")
            # sleep(10900)
            web.wait_element('//*[@id="sel_rep_accord"]/h3[1]/a')

            # # ? Send already filled forms
            # # ? Uncomment it if you want to update last saved form
            if web.wait_element('//*[@id="sel_rep_accord"]/h3[2]/a', timeout=1):
                web.execute_script_click_xpath("//*[@id='sel_rep_accord']/h3[2]/a")

                web.execute_script_click_js("body > div:nth-child(18) > div.ui-dialog-buttonpane.ui-widget-content.ui-helper-clearfix > div > button:nth-child(1)")

                web.wait_element("//a[contains(text(), 'Страница 1')]")

                sleep(5)
                site1, site2, site3 = 0, 0, 0
                fact1, fact2, fact3 = values_first_part[0], values_first_part[1], values_first_part[2]
                for ind, group in enumerate(groups):
                    # sleep(1)
                    if group == 'Объем розничной торговли':
                        try:
                            site1 = web.find_element(f"//*[contains(text(), '{group}')]/following-sibling::*[contains(@role, 'gridcell')][1]").get_attr("title")
                            print('OOOU EE:', web.find_element(f"//*[contains(text(), '{group}')]/following-sibling::*[contains(@role, 'gridcell')][1]").get_attr("title"))
                            # web.execute_script(xpath=f"//*[contains(text(), '{group}')]/following-sibling::*[contains(@role, 'gridcell')][1]", value=str(values_first_part[0]))
                        except:
                            # web.execute_script(xpath=f"//*[contains(text(), '{group}')]/following-sibling::*[contains(@role, 'gridcell')][1]", value='0')
                            pass

                        try:
                            site2 = web.find_element(f"//*[contains(text(), '{group}')]/following-sibling::*[contains(@role, 'gridcell')][2]").get_attr("title")
                            print('OOOU EE:', web.find_element(f"//*[contains(text(), '{group}')]/following-sibling::*[contains(@role, 'gridcell')][2]").get_attr("title"))
                            # web.execute_script(xpath=f"//*[contains(text(), '{group}')]/following-sibling::*[contains(@role, 'gridcell')][2]", value=str(values_first_part[1]))
                        except:
                            # web.execute_script(xpath=f"//*[contains(text(), '{group}')]/following-sibling::*[contains(@role, 'gridcell')][2]", value='0')
                            pass

                    elif group == 'Товарные запасы на конец отчетного месяца':
                        try:
                            site3 = web.find_element(f"//*[contains(text(), '{group}')]/following-sibling::*[contains(@role, 'gridcell')][1]").get_attr("title")
                            print('OOOU EE:', web.find_element(f"//*[contains(text(), '{group}')]/following-sibling::*[contains(@role, 'gridcell')][1]").get_attr("title"))
                            # web.execute_script(xpath=f"//*[contains(text(), '{group}')]/following-sibling::*[contains(@role, 'gridcell')][1]", value=str(values_first_part[2]))
                        except:
                            # web.execute_script(xpath=f"//*[contains(text(), '{group}')]/following-sibling::*[contains(@role, 'gridcell')][1]", value='0')
                            pass

                    # else:
                    #     try:
                    #         cur_val = round(round(values_second_part.get(group)) / 1000)
                    #         if cur_val < 100:
                    #             cur_val = 100
                    #         if cur_val > 9999:
                    #             cur_val = 9999
                    #         web.execute_script(xpath=f"//*[contains(text(), '{group}')]/following-sibling::*[contains(@role, 'gridcell')][2]", value=cur_val)
                    #     except:
                    #         web.execute_script(xpath=f"//*[contains(text(), '{group}')]/following-sibling::*[contains(@role, 'gridcell')][2]", value='0')

                if fact1 == int(site1) and fact2 == int(site2) and fact3 == int(site3):
                    insert_data_in_db(started_time=datetime.datetime.now(), store_id=store_id, store_name=store, status='success', error_reason='', error_saved_path='', execution_time='0', fact1=fact1, fact2=fact2, fact3=fact3, site1=site1, site2=site2, site3=site3)

                    web.close()
                    web.quit()
                    return ['success', '']
                else:
                    saved_path = save_screenshot(store)

                    insert_data_in_db(started_time=datetime.datetime.now(), store_id=store_id, store_name=store, status='FUCKING FAILED', error_reason='различия между фактом и на сайте', error_saved_path=saved_path, execution_time='0', fact1=fact1, fact2=fact2, fact3=fact3, site1=site1, site2=site2, site3=site3)

                    web.close()
                    web.quit()
                    return ['FUCKING FAILED', saved_path]
                # save_and_send(web, save=False)
                #
                # sign_ecp(ecp_sign)
                # # sleep(1000)
                # # print(store.replace('Торговый зал', '').replace(' ', '').replace('№', ''))
                # # sleep(30)
                #
                # wait_image_loaded()
                #
                # web.close()
                # web.quit()
                #
                # print('Successed')
                # return ['success', '']

            else:
                return ['failed', 'Нет сохранённого отчёта']

            # ? Open new report to fill it
            web.execute_script_click_js("body > div:nth-child(18) > div.ui-dialog-buttonpane.ui-widget-content.ui-helper-clearfix > div > button:nth-child(1)")

            web.wait_element("//a[contains(text(), 'Страница 1')]")
            web.find_element("//a[contains(text(), 'Страница 1')]").click()

            web.find_element('//*[@id="rtime"]').select('2')
            sleep(1)
            for ind, group in enumerate(groups):
                # sleep(1)
                if group == 'Объем розничной торговли':
                    try:
                        web.execute_script(xpath=f"//*[contains(text(), '{group}')]/following-sibling::*[contains(@role, 'gridcell')][1]", value=str(values_first_part[0]))
                        print('OOOU EE:', web.find_element(f"//*[contains(text(), '{group}')]/following-sibling::*[contains(@role, 'gridcell')][1]").get_attr("title"))
                    except:
                        web.execute_script(xpath=f"//*[contains(text(), '{group}')]/following-sibling::*[contains(@role, 'gridcell')][1]", value='0')

                    try:
                        web.execute_script(xpath=f"//*[contains(text(), '{group}')]/following-sibling::*[contains(@role, 'gridcell')][2]", value=str(values_first_part[1]))
                        print('OOOU EE:', web.find_element(f"//*[contains(text(), '{group}')]/following-sibling::*[contains(@role, 'gridcell')][2]").get_attr("title"))
                    except:
                        web.execute_script(xpath=f"//*[contains(text(), '{group}')]/following-sibling::*[contains(@role, 'gridcell')][2]", value='0')

                elif group == 'Товарные запасы на конец отчетного месяца':
                    try:
                        web.execute_script(xpath=f"//*[contains(text(), '{group}')]/following-sibling::*[contains(@role, 'gridcell')][1]", value=str(values_first_part[2]))
                        print('OOOU EE:', web.find_element(f"//*[contains(text(), '{group}')]/following-sibling::*[contains(@role, 'gridcell')][1]").get_attr("title"))
                    except:
                        web.execute_script(xpath=f"//*[contains(text(), '{group}')]/following-sibling::*[contains(@role, 'gridcell')][1]", value='0')

                else:
                    try:
                        cur_val = round(round(values_second_part.get(group)) / 1000)
                        if cur_val < 100:
                            cur_val = 100
                        if cur_val > 9999:
                            cur_val = 9999
                        web.execute_script(xpath=f"//*[contains(text(), '{group}')]/following-sibling::*[contains(@role, 'gridcell')][2]", value=str(cur_val))
                    except:
                        web.execute_script(xpath=f"//*[contains(text(), '{group}')]/following-sibling::*[contains(@role, 'gridcell')][2]", value='0')
            sleep(0.1)
            logger.info('Check4')
            web.find_element("//a[contains(text(), 'Данные исполнителя')]").click()
            web.execute_script(element_type="value", xpath="//*[@id='inpelem_1_0']", value='Қалдыбек Б.Ғ.')
            web.execute_script(element_type="value", xpath="//*[@id='inpelem_1_1']", value='87073332438')
            web.execute_script(element_type="value", xpath="//*[@id='inpelem_1_2']", value='87073332438')
            web.execute_script(element_type="value", xpath="//*[@id='inpelem_1_3']", value='KALDYBEK.B@magnum.kz')
            sleep(100)
            save_and_send(web, save=True)

            sign_ecp(ecp_sign)
            # sleep(1000)

            wait_image_loaded()

            web.close()
            web.quit()

            print('Successed')
            return ['success', '']

    else:
        web.close()
        web.quit()

        print('Srok istek')
        return ['failed', 'Срок ЭЦП истёк']


def get_data_from_1157(branch):

    for file in os.listdir(os.path.join(download_path, 'Splitted')):
        if branch + '_' in file:
            print('Read', os.path.join(os.path.join(download_path, 'Splitted'), file))
            return pd.read_excel(os.path.join(os.path.join(download_path, 'Splitted'), file), header=0)


if __name__ == '__main__':

    # create_and_send_final_report()
    # exit()

    # drop_table()

    sql_create_table()
    # start = datetime.datetime.now().strftime("%d.%m.%Y %H:%M:%S.%f")
    # insert_data_in_db(started_time=start, store_id=2350, store_name='Loh', status='failed', error_reason='guano', error_saved_path='', execution_time='10s')
    # exit()

    update_credentials(Path(r'\\vault.magnum.local\common\Stuff\_06_Бухгалтерия\! Актуальные ЭЦП'), owa_username, owa_password)

    all_branches = []

    counter = 2

    df = pd.DataFrame(columns=['id', 'branch', 'data'])
    print(download_path)
    for file in os.listdir(download_path):

        if file == '!result.xlsx':

            book = load_workbook(os.path.join(download_path, file))
            sheet = book.active
            # month = datetime.datetime.today().month
            month = 0

            months = 'DEFGHIJKLMNO'

            for ind, letter in enumerate(months):
                if sheet[letter + '2'].value is not None:
                    print(letter)
                    month = ind + 1

            print('MON:', month, months[month - 1])
            while sheet[f'A{counter}'].value is not None:
                all_branches.append([sheet[f'A{counter}'].value, sheet[f'B{counter}'].value, [sheet[f'{months[month - 1]}{counter}'].value, sheet[f'{months[month - 1]}{counter + 1}'].value, sheet[f'{months[month - 1]}{counter + 2}'].value]])

                row = pd.DataFrame({'id': sheet[f'A{counter}'].value, 'branch': sheet[f'B{counter}'].value, 'data': [[sheet[f'{months[month - 1]}{counter}'].value, sheet[f'{months[month - 1]}{counter + 1}'].value, sheet[f'{months[month - 1]}{counter + 2}'].value]]})
                df = pd.concat([df, row], ignore_index=True)
                counter += 3

    while True:
        try:
            df1 = get_all_branches_with_codes()
            break
        except:
            print('Error with adb')
            sleep(120)

    df['name'] = None
    df['store_id'] = None
    df['store_normal_name'] = None

    for i in range(len(df)):
        df['name'].loc[i] = df1[df1['branch_id'] == df['id'].iloc[i]]['store_name'].iloc[0]
        df['store_id'].loc[i] = df1[df1['branch_id'] == df['id'].iloc[i]]['store_id'].iloc[0]
        df['store_normal_name'].loc[i] = df1[df1['branch_id'] == df['id'].iloc[i]]['store_normal_name'].iloc[0]

    # ? Execute only rows with exact status

    status = 'succ1ess'

    # if status != 'success':
    #     all_rows = get_all_data()
    #
    #     all_bad_rows = all_rows[all_rows['status'] != 'success']
    #
    #     all_bad_rows['store_normal_name'] = None
    #     all_bad_rows = all_bad_rows.reset_index(inplace=False)
    #     for i in range(len(all_bad_rows)):
    #         all_bad_rows.loc[i, 'store_normal_name'] = df1[df1['store_id'] == all_bad_rows['store_id'].iloc[i]]['store_normal_name'].iloc[0]
    #
    #     df_prev = df.copy()
    #     df = all_bad_rows
    #     df['data'] = df_prev['data']

    print('Len:', len(df))
    check = False

    for ind, branch in enumerate(np.asarray(df['name'])):
        # print(df['store_normal_name'].iloc[ind], branch)
        # continue
        # if get_data_by_name(branch) != 0:
        #     continue
        # if 'Т' not in branch[1:]:
        #     continue
        # if branch != 'Торговый зал ШФ №4':
        #     continue
        # stores = ['Торговый зал АФ №51', 'Торговый зал АФ №21', 'Торговый зал АФ №16', 'Торговый зал ФКС №1', 'Торговый зал АФ №20', 'Торговый зал АСФ №12']
        # print(branch)
        # if branch in stores:
        #     continue
        # if branch == 'Торговый зал АСФ №70':
        #     check = True
        # if not check:
        #     print(branch)
        #     continue

        ecp_path = fr'\\vault.magnum.local\common\Stuff\_06_Бухгалтерия\! Актуальные ЭЦП\{branch}'

        if os.path.exists(ecp_path) and os.path.isdir(ecp_path):
            start = datetime.datetime.now().strftime("%d.%m.%Y %H:%M:%S.%f")
            start_time = time.time()
            try:
                print('Started', ind, branch)
                send_message_to_tg(tg_token, chat_id, f'Started, {ind}, {branch}')

                # ? Get the sum of each subgroup to fill
                data_from_1157 = get_data_from_1157(df[df['name'] == branch]['store_normal_name'].iloc[0])

                dick = dict()
                for j in data_from_1157['Подгруппа'].unique():
                    dick.update({j: sum(data_from_1157[data_from_1157['Подгруппа'] == j]['Фактические остатки'])})

                keys = list(dick.keys())

                # ? Setting values for each group with corresponding mapping
                for i in keys:
                    dick[dick1.get(i)] = dick.pop(i)

                print()
                print(df[df['name'] == branch])
                logger.info(df[df['name'] == branch]['data'].iloc[0])
                logger.info(dick)

                try:
                    status, saved_path = start_single_branch(ecp_path, branch, df[df['name'] == branch]['data'].iloc[0], dick, int(df['store_id'].iloc[ind]))
                except Exception as poebotnya:
                    logger.info(poebotnya)
                    # end_time = time.time()
                    # insert_data_in_db(started_time=start, store_id=int(df['store_id'].iloc[ind]), store_name=str(branch), status='poebotnya', error_reason=poebotnya, error_saved_path='', execution_time=str(end_time - start_time))
                    continue

                end_time = time.time()

                # if status != 'success':
                #     insert_data_in_db(started_time=start, store_id=int(df['store_id'].iloc[ind]), store_name=str(branch), status='failed', error_reason=status, error_saved_path=saved_path, execution_time=str(end_time - start_time))
                # else:
                #     insert_data_in_db(started_time=start, store_id=int(df['store_id'].iloc[ind]), store_name=str(branch), status='success', error_reason='', error_saved_path='', execution_time=str(end_time - start_time))

                send_message_to_tg(tg_token, chat_id, f'Finished, {ind}, {branch}')

            except Exception as ebanko:
                send_message_to_tg(tg_token, chat_id, f'Fucking error occured: {ebanko}')
                end_time = time.time()
                saved_path = save_screenshot(df['name'].iloc[ind])
                # insert_data_in_db(started_time=start, store_id=int(df['store_id'].iloc[ind]), store_name=str(branch), status='polomalsya', error_reason=str(ebanko), error_saved_path=saved_path, execution_time=str(end_time - start_time))
    # print(k)
    # create_and_send_final_report()
