import datetime
import os
import random
import shutil
import sys
import time
import uuid
from contextlib import suppress
from pathlib import Path
from time import sleep

import numpy as np
import openpyxl
import pandas as pd

import psycopg2 as psycopg2
from openpyxl import load_workbook
from pywinauto import keyboard

from config import logger, robot_name, db_host, db_port, db_name, db_user, db_pass, owa_username, owa_password, working_path, download_path, tg_token, chat_id, ip_address, saving_path, production_calendar, template_path, main_executor
from rpamini import Web, App
from tools import update_credentials, send_message_to_tg, send_file_to_tg
from pyautogui import screenshot

from openpyxl import load_workbook

from openpyxl.styles import PatternFill, Alignment

dick1 = {'РИС ВЕС': 'Рис',
         'КРУПА ГРЕЧНЕВАЯ ВЕС': 'гречневая',
         'МАСЛО ПОДСОЛНЕЧНОЕ': 'подсолнечн',
         'КАПУСТА': 'белокоч',
         'ЛУК, ЧЕСНОК': 'репчатый',
         'МОРКОВЬ': 'Морковь',
         'КАРТОФЕЛЬ': 'Картофель',
         'САХАР ПЕСОК': 'Сахар',
         'СОЛЬ ОБЫЧНАЯ': 'Соль'}

groups = ['Объем розничной торговли',
          'Товарные запасы на конец отчетного месяца',
          'Рис',
          'гречневая',
          'подсолнечн',
          'белокоч',
          'репчатый',
          'Морковь',
          'Картофель',
          'Сахар',
          'Соль']


def sql_drop_table():
    conn = psycopg2.connect(host=db_host, port=db_port, database=db_name, user=db_user, password=db_pass)
    table_create_query = f'''
            drop table robot.{robot_name.replace("-", "_")}
            '''
    c = conn.cursor()
    c.execute(table_create_query)

    conn.commit()
    c.close()
    conn.close()


def sql_create_table():
    conn = psycopg2.connect(host=db_host, port=db_port, database=db_name, user=db_user, password=db_pass)
    table_create_query = f'''
        CREATE TABLE IF NOT EXISTS ROBOT.{robot_name.replace("-", "_")} (
            started_time timestamp,
            ended_time timestamp,
            store_id int UNIQUE,
            store_name text UNIQUE,
            full_name text UNIQUE,
            executor_name text,
            status text,
            error_reason text,
            error_saved_path text,
            execution_time text,
            ecp_path text,
            fact1 text,
            fact2 text,
            fact3 text,
            site1 text,
            site2 text,
            site3 text
            )
        '''
    c = conn.cursor()
    c.execute(table_create_query)

    conn.commit()
    c.close()
    conn.close()


def delete_by_id(id):
    conn = psycopg2.connect(host=db_host, port=db_port, database=db_name, user=db_user, password=db_pass)
    table_create_query = f'''
                DELETE FROM ROBOT.{robot_name.replace("-", "_")} WHERE id = '{id}'
                '''
    c = conn.cursor()
    c.execute(table_create_query)
    conn.commit()
    c.close()
    conn.close()


def get_all_data():
    conn = psycopg2.connect(host=db_host, port=db_port, database=db_name, user=db_user, password=db_pass)
    table_create_query = f'''
            SELECT * FROM ROBOT.{robot_name.replace("-", "_")}
            order by started_time asc
            '''
    cur = conn.cursor()
    cur.execute(table_create_query)

    df1 = pd.DataFrame(cur.fetchall())
    df1.columns = ['started_time', 'ended_time', 'store_id', 'name', 'full_name', 'executor_name', 'status', 'error_reason', 'error_saved_path', 'execution_time', 'ecp_path', 'fact1', 'fact2', 'fact3', 'site1', 'site2', 'site3']

    cur.close()
    conn.close()

    return df1


def get_data_by_name(store_name):
    conn = psycopg2.connect(host=db_host, port=db_port, database=db_name, user=db_user, password=db_pass)
    table_create_query = f'''
            SELECT * FROM ROBOT.{robot_name.replace("-", "_")}
            where store_name = '{store_name}'
            order by started_time desc
            '''
    cur = conn.cursor()
    cur.execute(table_create_query)

    df1 = pd.DataFrame(cur.fetchall())
    # df1.columns = ['started_time', 'ended_time', 'store_id', 'name', 'status', 'error_reason', 'error_saved_path', 'execution_time']

    cur.close()
    conn.close()

    return len(df1)


def get_data_to_execute():
    conn = psycopg2.connect(host=db_host, port=db_port, database=db_name, user=db_user, password=db_pass)
    table_create_query = f'''
            SELECT * FROM ROBOT.{robot_name.replace("-", "_")}
            where (status = 'new' and (executor_name is NULL or executor_name = '{ip_address}'))
            --or (status = 'failed' and (executor_name is NULL or executor_name = '{ip_address}'))
            or (status = 'processing' and (executor_name is NULL or executor_name = '{ip_address}'))
            order by started_time desc
            '''
    cur = conn.cursor()
    cur.execute(table_create_query)

    df1 = pd.DataFrame(cur.fetchall())

    with suppress(Exception):
        df1.columns = ['started_time', 'ended_time', 'store_id', 'name', 'full_name', 'executor_name', 'status', 'error_reason', 'error_saved_path', 'execution_time', 'ecp_path', 'fact1', 'fact2', 'fact3', 'site1', 'site2', 'site3']

    for ind1 in range(len(df1)):
        df1.loc[ind1, 'name'] = str(df1['name'].iloc[ind1]).replace('_ОПТ', '')

    cur.close()
    conn.close()

    return df1


def insert_data_in_db(started_time, store_id, store_name, full_name, executor_name, status_, error_reason, error_saved_path, execution_time, ecp_path_, fact1, fact2, fact3, site1, site2, site3):
    conn = psycopg2.connect(host=db_host, port=db_port, database=db_name, user=db_user, password=db_pass)

    print('Started inserting')
    # query_delete_id = f"""
    #         delete from ROBOT.{robot_name.replace("-", "_")}_2 where store_id = '{store_id}'
    #     """
    query_delete = f"""
        delete from ROBOT.{robot_name.replace("-", "_")} where store_name = '{store_name}'
    """
    query = f"""
        INSERT INTO ROBOT.{robot_name.replace("-", "_")} (started_time, ended_time, store_id, store_name, full_name, executor_name, status, error_reason, error_saved_path, execution_time, ecp_path, fact1, fact2, fact3, site1, site2, site3)
        VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
    """
    # ended_time = '' if status_ != 'success' else datetime.datetime.now().strftime("%d.%m.%Y %H:%M:%S.%f")
    ended_time = datetime.datetime.now().strftime("%d.%m.%Y %H:%M:%S.%f")
    values = (
        started_time,
        ended_time,
        str(store_id),
        store_name,
        full_name,
        executor_name,
        status_,
        error_reason,
        error_saved_path,
        execution_time,
        ecp_path_,
        str(fact1),
        str(fact2),
        str(fact3),
        site1,
        site2,
        site3
    )

    print(values)

    cursor = conn.cursor()

    cursor.execute(query_delete)
    # conn.autocommit = True
    try:
        cursor.execute(query_delete)
        # cursor.execute(query_delete_id)
    except Exception as e:
        print('GOVNO', e)
        pass
    try:
        cursor.execute(query, values)
    except Exception as e:
        conn.rollback()
        print(f"Error: {e}")

    conn.commit()

    cursor.close()
    conn.close()


def get_all_branches_with_codes():
    import psycopg2
    import pandas as pd

    conn = psycopg2.connect(dbname='adb', host='172.16.10.22', port='5432',
                            user='rpa_robot', password='Qaz123123+')

    cur = conn.cursor(name='1583_first_part')

    query = f"""
        select db.id_sale_object, ds.source_store_id, ds.store_name, ds.sale_obj_name 
        from dwh_data.dim_branches db
        left join dwh_data.dim_store ds on db.id_sale_object = ds.sale_source_obj_id
        where ds.store_name like '%Торговый%' and current_date between ds.datestart and ds.dateend
        group by db.id_sale_object, ds.source_store_id, ds.store_name, ds.sale_obj_name
        order by ds.source_store_id
    """

    cur.execute(query)

    print('Executed')

    df1 = pd.DataFrame(cur.fetchall())
    df1.columns = ['branch_id', 'store_id', 'store_name', 'store_normal_name']

    cur.close()
    conn.close()

    return df1


def sign_ecp(ecp):
    logger.info('Started ECP')
    # print('Kek')
    logger.info(f'KEY: {ecp}')
    # print(f'KEY: {ecp}')
    app = App('')

    el = {"title": "Открыть файл", "class_name": "SunAwtDialog", "control_type": "Window",
          "visible_only": True, "enabled_only": True, "found_index": 0, "parent": None}

    if app.wait_element(el, timeout=30):

        keyboard.send_keys(ecp.replace('(', '{(}').replace(')', '{)}'), pause=0.01, with_spaces=True)
        sleep(0.05)
        keyboard.send_keys('{ENTER}')

        if app.wait_element({"title_re": "Формирование ЭЦП.*", "class_name": "SunAwtDialog", "control_type": "Window",
                             "visible_only": True, "enabled_only": True, "found_index": 0, "parent": None}, timeout=30):
            app.find_element({"title_re": "Формирование ЭЦП.*", "class_name": "SunAwtDialog", "control_type": "Window",
                              "visible_only": True, "enabled_only": True, "found_index": 0, "parent": None}).type_keys('Aa123456')

            # keyboard.send_keys('Aa123456')
            sleep(2)

            keyboard.send_keys('{ENTER}')
            sleep(3)

            keyboard.send_keys('{ENTER}')
            app = None
            logger.info('Finished ECP')
        else:
            logger.info('Quit mazafaka1')
            app = None
            return 'broke'
    else:
        logger.info('Quit mazafaka')
        # try:
        #     app.find_element({"title": "Закрыть", "class_name": "", "control_type": "Button",
        #                       "visible_only": True, "enabled_only": True, "found_index": 0}, ti).click()
        #     app.quit()
        # except:
        #     pass
        app = None
        return 'broke'
    # app.quit()


def save_screenshot(store):
    scr = screenshot()
    scr_path = str(os.path.join(os.path.join(saving_path, 'Ошибки 2т'), str(store + '.png')))
    scr.save(scr_path)

    return scr_path


def wait_loading(web, xpath):
    print('Started loading')
    ind = 0
    element = ''
    while True:
        try:
            print(web.lolus('//*[@id="loadmask-1315"]'))
            if web.lolus('//*[@id="loadmask-1315"]') == '':
                element = ''
            if (element == '' and web.lolus('//*[@id="loadmask-1315"]') == 'none') or (ind >= 500):
                print('Loaded')
                sleep(0.5)
                break
        except:
            print('No loader')
            break
        ind += 1
        sleep(0.05)


def create_and_send_final_report():

    df = get_all_data()

    df.columns = ['Время начала', 'Время окончания', 'Номер филиала', 'Название филиала', 'Полное название', 'Машина', 'Статус', 'Причина ошибки', 'Пусть сохранения скриншота', 'Время исполнения (сек)', 'Путь к ЭЦП', 'Факт1', 'Факт2', 'Факт3', 'Сайт1', 'Сайт2', 'Сайт3']

    with suppress(Exception):
        df['Время исполнения (сек)'] = df['Время исполнения (сек)'].astype(float)
        df['Время исполнения (сек)'] = df['Время исполнения (сек)'].round()

    df.to_excel('result.xlsx', index=False)

    workbook = load_workbook('result.xlsx')
    sheet = workbook.active

    red_fill = PatternFill(start_color="FFA864", end_color="FFA864", fill_type="solid")
    green_fill = PatternFill(start_color="A6FF64", end_color="A6FF64", fill_type="solid")

    for cell in sheet['G']:
        if cell.value == 'failed':
            cell.fill = red_fill
        if cell.value == 'success':
            cell.fill = green_fill

    for col in 'ABCDGH':

        max_length = max(len(str(cell.value)) for cell in sheet[col])

        if col == 'A' or col == 'B':
            max_length -= 3
        if col == 'D':
            max_length += 5
        if col == 'A':
            max_length -= 3

        sheet.column_dimensions[col].width = max_length

    for col in 'ABCDGEFGH':
        for cell in sheet[col]:
            cell.alignment = Alignment(horizontal='center')

    workbook.save('result.xlsx')

    # send_file_to_tg(tg_token, chat_id, 'Отправляем отчёт по заполнению', 'result.xlsx')


def wait_image_loaded():
    found = False
    for i in range(230):
        for file in os.listdir(download_path):
            if '.jpg' in file and 'crdownload' not in file and (time.time() - os.path.getctime(os.path.join(download_path, file))) <= 100:
                shutil.move(os.path.join(download_path, file), os.path.join(os.path.join(saving_path, 'Отчёты 2т'), branch + '.jpg'))
                print(file)
                print(os.path.join(os.path.join(saving_path, 'Отчёты 2т'), branch + '.jpg'))
                print('---')
                found = True
                break
        if found:
            break
        sleep(1)

    return found


def save_and_send(web, ecp, save):
    print('Saving and Sending')
    if save:
        web.execute_script_click_xpath("//span[text() = 'Сохранить']")
        sleep(1)
        print('Clicked Save')
        if web.wait_element("//span[text() = 'Сохранить отчет и Удалить другие']", timeout=5):
            web.execute_script_click_xpath("//span[text() = 'Сохранить отчет и Удалить другие']")
    print('Clicking Send')
    web.execute_script_click_xpath("//span[text() = 'Отправить']")
    print('Clicked Send')
    web.wait_element("//input[@value = 'Персональный компьютер']", timeout=30)
    web.execute_script_click_xpath("//input[@value = 'Персональный компьютер']")

    sign_ecp(ecp)

    if web.wait_element("//h1[contains(text(), 'Whitelabel')]", timeout=5):
        for _ in range(10):
            try:
                web.execute_script_click_xpath("//span[text() = 'Сохранить']")
            except:
                sleep(60)
                web.execute_script_click_xpath("//span[text() = 'Сохранить']")
            sleep(1)
            if web.wait_element("//span[text() = 'Сохранить отчет и Удалить другие']", timeout=30):
                web.execute_script_click_xpath("//span[text() = 'Сохранить отчет и Удалить другие']")
            web.execute_script_click_xpath("//button[@class='btn-savesigned ui-button ui-widget ui-state-default ui-corner-all ui-button-text-icon-primary']/span[text() = 'Отправить']")

            if web.wait_element("//input[@value = 'Персональный компьютер']", timeout=60):
                web.execute_script_click_xpath("//input[@value = 'Персональный компьютер']")
            else:
                web.find_element("//button[@class='btn-savesigned ui-button ui-widget ui-state-default ui-corner-all ui-button-text-icon-primary']/span[text() = 'Отправить']").click()
                web.wait_element("//input[@value = 'Персональный компьютер']", timeout=120)
                web.execute_script_click_xpath("//input[@value = 'Персональный компьютер']")
            print('Checkpoint on signing')
            sign_ecp(ecp)

            if web.wait_element("//span[text() = 'Продолжить']", timeout=10):
                web.execute_script_click_xpath("//span[text() = 'Продолжить']")

            if not web.wait_element("//h1[contains(text(), 'Whitelabel')]", timeout=5):
                break


def proverka_ecp(web):

    if web.wait_element('//*[@id="AgreeId_header_hd-textEl"]', timeout=.5):
        web.execute_script_click_xpath("//span[text() = 'Согласен']")


def start_single_branch(filepath, store, values_first_part, values_second_part):

    print('Started web')

    if any(val < 0 for val in values_first_part if val is not None):
        return ['failed', 'Отрицательная сумма в первой части 2Т', [None, None, None]]

    web = Web()
    web.run()
    web.get('https://cabinet.stat.gov.kz/')
    logger.info('Check-1')

    proverka_ecp(web=web)

    web.wait_element('//*[@id="idLogin"]')
    web.find_element('//*[@id="idLogin"]').click()

    proverka_ecp(web=web)

    # web.wait_element('//*[@id="button-1077-btnEl"]')
    # web.find_element('//*[@id="button-1077-btnEl"]').click()

    web.wait_element('//*[@id="lawAlertCheck"]')
    web.find_element('//*[@id="lawAlertCheck"]').click()

    sleep(0.5)
    web.find_element('//*[@id="loginButton"]').click()

    logger.info('Check-2')
    ecp_auth = ''
    ecp_sign = ''
    for files in os.listdir(filepath):
        if 'AUTH' in files:
            ecp_auth = os.path.join(filepath, files)
        if 'GOST' in files:
            ecp_sign = os.path.join(filepath, files)

    sleep(1)
    sign_ecp(ecp_auth)

    logged_in = web.wait_element('//*[@id="idLogout"]/a', timeout=60)
    # sleep(1000)
    if logged_in:
        if web.find_element("//a[text() = 'Выйти']"):

            if web.wait_element("//span[contains(text(), 'Пройти позже')]", timeout=5):
                try:
                    web.find_element("//span[contains(text(), 'Пройти позже')]").click()
                except:
                    save_screenshot(store)
            logger.info('Check0')
            if web.wait_element('//*[@id="dontAgreeId-inputEl"]', timeout=5):
                web.find_element('//*[@id="dontAgreeId-inputEl"]').click()
                sleep(0.3)
                web.find_element('//*[@id="saveId-btnIconEl"]').click()
                sleep(1)

                # * --- Deprecated (maybe useful)
                # web.find_element('//*[@id="ext-gen1893"]').click()
                # web.find_element('//*[@id="boundlist-1327-listEl"]/ul/li').click()
                # * ---

                web.wait_element('//*[@id="keyCombo-inputEl"]')

                web.execute_script_click_xpath("//*[@id='keyCombo-inputEl']/../following-sibling::td//div")

                web.find_element("//li[contains(text(), 'Персональный компьютер')]").click()
                sleep(1.5)

                web.execute_script_click_xpath("//span[contains(text(), 'Продолжить')]")

                print('Done lol')
                sign_ecp(ecp_sign)
                print()
                try:
                    web.wait_element("//span[contains(text(), 'Пройти позже')]", timeout=5)
                    web.find_element("//span[contains(text(), 'Пройти позже')]").click()

                except:
                    pass
            logger.info('Check1')

            web.wait_element("//span[contains(text(), 'Мои отчёты')]")
            web.execute_script_click_xpath("//span[contains(text(), 'Мои отчёты')]")

            # sleep(0.7)

            # web.wait_element('//*[@id="radio-1131-boxLabelEl"]')

            if web.wait_element("//span[contains(text(), 'Пройти позже')]", timeout=1.5):
                web.execute_script_click_xpath("//span[contains(text(), 'Пройти позже')]")
            sleep(1)

            # ? Check if 2T exists
            # for i in range(1):
            #
            # wait_loading(web, '//*[@id="loadmask-1315"]')
            # if web.wait_element("//span[contains(text(), 'Пройти позже')]", timeout=1.5):
            #     web.execute_script_click_xpath("//span[contains(text(), 'Пройти позже')]")
            # if web.wait_element("//div[contains(text(), '2-торговля')]", timeout=3):
            #     web.find_element("//div[contains(text(), '2-торговля')]").click()
            # else:
            #     if i < 3:

            if web.wait_element("//span[contains(text(), 'Пройти позже')]", timeout=1.5):
                web.execute_script_click_xpath("//span[contains(text(), 'Пройти позже')]")

            for _ in range(3):

                is_loaded = True if len(web.find_elements("//div[contains(@class, 'x-grid-row-expander')]", timeout=15)) >= 1 else False

                if is_loaded:
                    if web.wait_element("//div[contains(text(), '2-торговля')]", timeout=3):
                        web.find_element("//div[contains(text(), '2-торговля')]").click()

                    else:
                        saved_path = save_screenshot(store)
                        web.close()
                        web.quit()

                        print('Return those shit')
                        return ['Нет 2-т', saved_path, '']

                else:
                    web.refresh()

            if web.wait_element("//span[contains(text(), 'Пройти позже')]", timeout=1.5):
                web.execute_script_click_xpath("//span[contains(text(), 'Пройти позже')]")
            # web.find_element('//*[@id="radio-1133-boxLabelEl"]').click()
            # wait_loading(web, '//*[@id="loadmask-1315"]')
            # web.refresh()

            sleep(0.5)

            web.find_element('//*[@id="createReportId-btnIconEl"]').click()

            sleep(1)

            # ? Switch to the opened tab
            web.driver.switch_to.window(web.driver.window_handles[-1])

            web.wait_element('//*[@id="td_select_period_level_1"]/span')
            web.execute_script_click_js("#btn-opendata")
            sleep(0.3)

            if web.lolus('/html/body/div[7]') == 'block':
                web.find_element('/html/body/div[7]/div[11]/div/button[2]').click()

                saved_path = save_screenshot(store)
                web.close()
                web.quit()

                print('Return that shit')
                return ['Выскочила херня', saved_path, '']

            web.wait_element('//*[@id="sel_statcode_accord"]/div/p/b[1]', timeout=100)
            web.execute_script_click_js("body > div:nth-child(16) > div.ui-dialog-buttonpane.ui-widget-content.ui-helper-clearfix > div > button:nth-child(1) > span")
            # sleep(10900)
            web.wait_element('//*[@id="sel_rep_accord"]/h3[1]/a')

            sites = []
            # # ? Send already filled forms
            # # ? Uncomment it if you want to update last saved form
            # if web.wait_element('//*[@id="sel_rep_accord"]/h3[2]/a', timeout=5):
            #     web.execute_script_click_xpath("//*[@id='sel_rep_accord']/h3[2]/a")
            #     print('Clicking')
            #     web.execute_script_click_js("body > div:nth-child(18) > div.ui-dialog-buttonpane.ui-widget-content.ui-helper-clearfix > div > button:nth-child(1)")
            #
            #     web.wait_element("//a[contains(text(), 'Страница 1')]", timeout=120)
            #
            #     sleep(.5)
            #     logger.info(values_first_part)
            #     for ind, group in enumerate(groups):
            #
            #         if group == 'Объем розничной торговли':
            #             try:
            #                 sites.append(int(web.find_element(f"//*[contains(text(), '{group}')]/following-sibling::*[contains(@role, 'gridcell')][1]").get_attr("title")))
            #             except:
            #                 sites.append(web.find_element(f"//*[contains(text(), '{group}')]/following-sibling::*[contains(@role, 'gridcell')][1]").get_attr("title"))
            #             web.execute_script(xpath=f"//*[contains(text(), '{group}')]/following-sibling::*[contains(@role, 'gridcell')][1]", value=str(values_first_part[0]))
            #
            #             try:
            #                 sites.append(int(web.find_element(f"//*[contains(text(), '{group}')]/following-sibling::*[contains(@role, 'gridcell')][2]").get_attr("title")))
            #             except:
            #                 sites.append(web.find_element(f"//*[contains(text(), '{group}')]/following-sibling::*[contains(@role, 'gridcell')][2]").get_attr("title"))
            #             web.execute_script(xpath=f"//*[contains(text(), '{group}')]/following-sibling::*[contains(@role, 'gridcell')][2]", value=str(values_first_part[1]))
            #
            #         elif group == 'Товарные запасы на конец отчетного месяца':
            #             try:
            #                 sites.append(int(web.find_element(f"//*[contains(text(), '{group}')]/following-sibling::*[contains(@role, 'gridcell')][1]").get_attr("title")))
            #             except:
            #                 sites.append(web.find_element(f"//*[contains(text(), '{group}')]/following-sibling::*[contains(@role, 'gridcell')][1]").get_attr("title"))
            #             web.execute_script(xpath=f"//*[contains(text(), '{group}')]/following-sibling::*[contains(@role, 'gridcell')][1]", value=str(values_first_part[2]))
            #
            #         # ? Filling second part of the 1st page
            #         else:
            #             try:
            #                 cur_val = round(round(values_second_part.get(group)) / 1)
            #             except:
            #                 print('ERROR:', values_second_part.get(group))
            #                 cur_val = 10
            #             if cur_val < 10:
            #                 cur_val = 10
            #             if cur_val > 9999:
            #                 cur_val = 9999
            #             logger.info(f'cur_val: {cur_val}')
            #             web.execute_script(xpath=f"//*[contains(text(), '{group}')]/following-sibling::*[contains(@role, 'gridcell')][2]", value=cur_val)
            #     sleep(0.1)
            #
            #     save_and_send(web, ecp=ecp_sign, save=False)
            #
            #     # sign_ecp(ecp_sign)
            #     # sleep(1000)
            #     # print(store.replace('Торговый зал', '').replace(' ', '').replace('№', ''))
            #     # sleep(30)
            #
            #     wait_image_loaded()
            #
            # web.close()
            # web.quit()
            #
            # print('Successed')
            #
            # if sites[0] == values_first_part[0] and sites[1] == values_first_part[1] and sites[2] == values_first_part[2]:
            #     return ['success', '', sites]
            # else:
            #     return ['success', 'Были разные данные', sites]

            # ? Open new report to fill it

            print('Clicking1')
            # web.execute_script_click_js("body > div:nth-child(18) > div.ui-dialog-buttonpane.ui-widget-content.ui-helper-clearfix > div > button:nth-child(1)")
            web.execute_script_click_xpath('/html/body/div[17]/div[11]/div/button[1]/span')
            web.wait_element("//a[contains(text(), 'Страница 1')]", timeout=10)
            web.find_element("//a[contains(text(), 'Страница 1')]").click()

            web.find_element('//*[@id="rtime"]').select('2')
            sleep(1)
            for ind, group in enumerate(groups):

                if group == 'Объем розничной торговли':
                    try:
                        sites.append(int(web.find_element(f"//*[contains(text(), '{group}')]/following-sibling::*[contains(@role, 'gridcell')][1]").get_attr("title")))
                    except:
                        sites.append(web.find_element(f"//*[contains(text(), '{group}')]/following-sibling::*[contains(@role, 'gridcell')][1]").get_attr("title"))
                    web.execute_script(xpath=f"//*[contains(text(), '{group}')]/following-sibling::*[contains(@role, 'gridcell')][1]",
                                       value=str(values_first_part[0]) if values_first_part[0] is not None else '0')

                    try:
                        sites.append(int(web.find_element(f"//*[contains(text(), '{group}')]/following-sibling::*[contains(@role, 'gridcell')][2]").get_attr("title")))
                    except:
                        sites.append(web.find_element(f"//*[contains(text(), '{group}')]/following-sibling::*[contains(@role, 'gridcell')][2]").get_attr("title"))
                    web.execute_script(xpath=f"//*[contains(text(), '{group}')]/following-sibling::*[contains(@role, 'gridcell')][2]",
                                       value=str(values_first_part[1]) if values_first_part[1] is not None else '0')

                elif group == 'Товарные запасы на конец отчетного месяца':
                    try:
                        sites.append(int(web.find_element(f"//*[contains(text(), '{group}')]/following-sibling::*[contains(@role, 'gridcell')][1]").get_attr("title")))
                    except:
                        sites.append(web.find_element(f"//*[contains(text(), '{group}')]/following-sibling::*[contains(@role, 'gridcell')][1]").get_attr("title"))
                    web.execute_script(xpath=f"//*[contains(text(), '{group}')]/following-sibling::*[contains(@role, 'gridcell')][1]",
                                       value=str(values_first_part[2]) if values_first_part[2] is not None else '0')

                # ? Filling second part of the 1st page
                else:
                    # cur_val = round(values_second_part.get(group))
                    try:
                        cur_val = round(round(values_second_part.get(group)) / 1)
                    except:
                        print('ERROR:', values_second_part.get(group))
                        cur_val = 10
                    if cur_val < 10:
                        cur_val = 10
                    if cur_val > 9999:
                        cur_val = 9999
                    logger.info(f'cur_val: {cur_val}')
                    web.execute_script(xpath=f"//*[contains(text(), '{group}')]/following-sibling::*[contains(@role, 'gridcell')][2]", value=cur_val)

            sleep(0.1)
            # ? Second page
            web.find_element("//a[contains(text(), 'Данные исполнителя')]").click()
            web.execute_script(element_type="value", xpath="//*[@id='inpelem_1_0']", value='Қалдыбек Б.Ғ.')
            web.execute_script(element_type="value", xpath="//*[@id='inpelem_1_1']", value='87073332438')
            web.execute_script(element_type="value", xpath="//*[@id='inpelem_1_2']", value='87073332438')
            web.execute_script(element_type="value", xpath="//*[@id='inpelem_1_3']", value='KALDYBEK.B@magnum.kz')
            sleep(0)

            for tries in range(100):

                save_and_send(web, ecp=ecp_sign, save=True)

                # sign_ecp(ecp_sign)
                # sleep(1000)

                found = wait_image_loaded()

                if found:

                    web.close()
                    web.quit()

                    print('Successed')

                    if sites[0] == values_first_part[0] and sites[1] == values_first_part[1] and sites[2] == values_first_part[2]:
                        return ['success', '', sites]
                    else:
                        return ['success', 'Были разные данные, робот изменил', sites]

                # return ['success', '', sites]

    else:

        saved_path = save_screenshot(store)

        web.close()
        web.quit()

        print('Srok istek')
        return ['Срок ЭЦП истёк', saved_path, '']


def get_data_from_1157(branch):
    logger.info('----------------------------')
    logger.info(branch)
    for file in os.listdir(r'\\172.16.8.87\d\.rpa\.agent\robot-1157-DWH\Output\Splitted'):
        if branch + '_' in file:
            logger.info('Read' + os.path.join(r'\\172.16.8.87\d\.rpa\.agent\robot-1157-DWH\Output\Splitted', file))
            logger.info('----------------------------')
            return pd.read_excel(os.path.join(r'\\172.16.8.87\d\.rpa\.agent\robot-1157-DWH\Output\Splitted', file), header=0)


def is_today_start():

    update_credentials(template_path, owa_username, owa_password)

    calendar = pd.read_excel(production_calendar)

    today_ = datetime.datetime.now().strftime('%d.%m.%y')

    cur_day_index = calendar[calendar['Day'] == today_]['Type'].index[0]
    cur_day_type = calendar[calendar['Day'] == today_]['Type'].iloc[0]

    count = 0
    day_ = None
    found = False

    for i in range(1, 31):

        try:
            day = int(calendar['Day'].iloc[cur_day_index + i].split('.')[0])
            print(calendar['Day'].iloc[cur_day_index + i], calendar['Weekday'].iloc[cur_day_index + i], calendar['Type'].iloc[cur_day_index + i])

        except:
            day = 1

        if day == 1:

            for j in range(1, 6):
                print(cur_day_index, i, j, cur_day_index + i - j)

                print('---', calendar['Day'].iloc[cur_day_index + i - j], calendar['Weekday'].iloc[cur_day_index + i - j], calendar['Type'].iloc[cur_day_index + i - j])

                if calendar['Type'].iloc[cur_day_index + i - j] == 'Working':
                    count += 1
                if count == 3:
                    found = True
                    day_ = calendar['Day'].iloc[cur_day_index + i - j]
                    break
        if found:
            break

    print(cur_day_index, cur_day_type)

    print(day_)

    if today_ == day_:  # * datetime.datetime.today().strftime('%d.%m.%y') == day_:
        return True
    else:
        return False


if __name__ == '__main__':

    if not is_today_start():
        logger.info(f'Not working day - {datetime.date.today()}')
        # exit()
    # create_and_send_final_report()
    # exit()

    # drop_table()

    print(ip_address)

    sql_create_table()
    # start = datetime.datetime.now().strftime("%d.%m.%Y %H:%M:%S.%f")
    # insert_data_in_db(started_time=start, store_id=2350, store_name='Loh', status='failed', error_reason='guano', error_saved_path='', execution_time='10s')
    # exit()
    # insert_data_in_db(started_time=start, store_name=str(branch), executor_name=ip_address, status_='processing', error_reason='', error_saved_path='', execution_time='', ecp_path_='', fact1='', fact2='', fact3='', site1='', site2='', site3='')
    update_credentials(Path(r'\\172.16.8.87\d\.rpa'), owa_username, owa_password)
    update_credentials(Path(r'\\vault.magnum.local\common\Stuff\_06_Бухгалтерия\! Актуальные ЭЦП'), owa_username, owa_password)

    all_branches = []

    counter = 2

    df = pd.DataFrame(columns=['id', 'branch', 'data'])

    for file in os.listdir(r'\\172.16.8.87\d\.rpa\.agent\robot-2t-dwh\Output\Выргузка 2Т'):

        if file == '!result.xlsx':

            book = load_workbook(os.path.join(r'\\172.16.8.87\d\.rpa\.agent\robot-2t-dwh\Output\Выргузка 2Т', file))
            sheet = book.active
            # month = datetime.datetime.today().month
            month = 0

            months = 'DEFGHIJKLMNO'

            for ind, letter in enumerate(months):
                if sheet[letter + '2'].value is not None:
                    print(letter)
                    month = ind + 1

            print('MON:', month, months[month - 1])
            while sheet[f'A{counter}'].value is not None:
                store_id = sheet[f'A{counter}'].value
                store_name = sheet[f'B{counter}'].value
                first, second, third = sheet[f'{months[month - 1]}{counter}'].value, sheet[f'{months[month - 1]}{counter + 1}'].value, sheet[f'{months[month - 1]}{counter + 2}'].value

                # insert_data_in_db(started_time='', store_id=store_id, store_name=store_name, full_name=store_name, executor_name=ip_address, status_='new', error_reason='', error_saved_path='', execution_time='', ecp_path_='', fact1='', fact2='', fact3='', site1='', site2='', site3='')

                all_branches.append([sheet[f'A{counter}'].value, sheet[f'B{counter}'].value, [sheet[f'{months[month - 1]}{counter}'].value, sheet[f'{months[month - 1]}{counter + 1}'].value, sheet[f'{months[month - 1]}{counter + 2}'].value]])

                row = pd.DataFrame({'id': sheet[f'A{counter}'].value, 'branch': sheet[f'B{counter}'].value, 'data': [[sheet[f'{months[month - 1]}{counter}'].value, sheet[f'{months[month - 1]}{counter + 1}'].value, sheet[f'{months[month - 1]}{counter + 2}'].value]]})
                df = pd.concat([df, row], ignore_index=True)
                counter += 3

    while True:
        try:
            df1 = get_all_branches_with_codes()
            break
        except:
            print('Error with adb dwh')
            sleep(120)

    df['name'] = None
    df['store_id'] = None
    df['store_normal_name'] = None

    for i in range(len(df)):
        df['name'].loc[i] = df1[df1['branch_id'] == df['id'].iloc[i]]['store_name'].iloc[0]
        df['store_id'].loc[i] = df1[df1['branch_id'] == df['id'].iloc[i]]['store_id'].iloc[0]
        df['store_normal_name'].loc[i] = df1[df1['branch_id'] == df['id'].iloc[i]]['store_normal_name'].iloc[0]

    # ? Execute only rows with exact status

    # status = 'success'

    # ? Dispatcher
    if ip_address == main_executor:

        sql_drop_table()

        sql_create_table()

        for ind in range(len(df)):
            row = df.iloc[ind]

            ecp_path = fr"\\vault.magnum.local\common\Stuff\_06_Бухгалтерия\! Актуальные ЭЦП\{row['name']}"
            # print(row)

            first_int = int(row['data'][0]) if row['data'][0] is not None else 0
            second_int = int(row['data'][1]) if row['data'][1] is not None else 0
            third_int = int(row['data'][2]) if row['data'][2] is not None else 0

            insert_data_in_db(started_time=datetime.datetime.now().strftime("%d.%m.%Y %H:%M:%S.%f"), store_id=int(row['id']), store_name=row['name'], full_name=row['branch'], executor_name=None, status_='new', error_reason='', error_saved_path='', execution_time='', ecp_path_=ecp_path, fact1=first_int, fact2=second_int, fact3=third_int, site1='', site2='', site3='')
    # exit()
    # if status != 'success':
    #     all_rows = get_all_data()
    #
    #     all_bad_rows = all_rows[all_rows['status'] != 'success']
    #
    #     all_bad_rows['store_normal_name'] = None
    #     all_bad_rows = all_bad_rows.reset_index(inplace=False)
    #
    #     for i in range(len(all_bad_rows)):
    #         all_bad_rows.loc[i, 'store_normal_name'] = df1[df1['store_id'] == all_bad_rows['store_id'].iloc[i]]['store_normal_name'].iloc[0]
    #
    #     df_prev = df.copy()
    #     df = all_bad_rows
    #     df['data'] = df_prev['data']

    print('Len:', len(df))
    check = False

    # ? Performer

    c = 0
    a = set()
    # create_and_send_final_report()
    # exit()
    status = None
    for ind in range(len(df)):

        all_data = get_data_to_execute()
        branch = df['name'].iloc[ind]
        full_name = df['branch'].iloc[ind]
        id_ = df['id'].iloc[ind]
        data = df['data'].iloc[ind]

        all_data = get_data_to_execute()

        data_to_execute = list(all_data['name'])

        logger.warning(f'Started1 {ind} | {branch}')

        if branch not in data_to_execute:
            print(f'Skipped {branch}')
            continue

        branch_in_1157 = branch
        # print(branch.replace('_ОПТ', ''), '|', data_to_execute)

        # if branch.replace('_ОПТ', '') not in data_to_execute:
        #     continue

        if branch == 'Торговый зал СТМ 1АФ':
            branch = 'РЦ DAMU Алматы'
            branch_in_1157 = 'РЦ №3 филиал в г.Алматы'

        skipping = False

        for file in os.listdir(os.path.join(saving_path, 'Отчёты 2т')):
            if branch == file.split('.')[0]:
                skipping = True
                break

        if skipping:
            continue

        # if f"{branch.split()[-2]} {branch.split()[-1]}" == 'КФ №4':
        #     continue

        # if f"{branch.split()[-2]} {branch.split()[-1]}" not in ['АФ №82']:
        #     print('EXEC1:', f"{branch.split()[-2]} {branch.split()[-1]}")
        #     continue
        # print(f"EXECUTING: {branch.split()[-2]} {branch.split()[-1]}")

        c += 1
        # continue
        # if 'АСФ №24' not in branch:
        #     continue
        a.update([branch])
        # continue
        ecp_path = fr'\\vault.magnum.local\common\Stuff\_06_Бухгалтерия\! Актуальные ЭЦП\{branch.replace("_ОПТ", "")}'
        # print('KEKUSSSSSSSSSSSSSSSSS', os.path.exists(ecp_path), os.path.isdir(ecp_path), os.path.exists(ecp_path) and os.path.isdir(ecp_path))
        if os.path.exists(ecp_path) and os.path.isdir(ecp_path):

            start = datetime.datetime.now().strftime("%d.%m.%Y %H:%M:%S.%f")
            start_time = time.time()
            try:
                print('Started', ind, branch)
                logger.warning(f'Started {ind} | {branch}')
                # send_message_to_tg(tg_token, chat_id, f'Начал, {ind}, {branch}')

                # ? Get the sum of each subgroup to fill
                data_from_1157 = get_data_from_1157(df[df['name'] == branch]['store_normal_name'].iloc[0])
                # print(data_from_1157)
                dick = dict()

                # ? Getting data from 1157 and adding it to the corresponding branch
                for j in data_from_1157['Подгруппа'].unique():
                    dick.update({j: sum(data_from_1157[data_from_1157['Подгруппа'] == j]['Фактические остатки'])})

                keys = list(dick.keys())

                # ? Setting values for each group with corresponding mapping
                for i in keys:
                    dick[dick1.get(i)] = dick.pop(i)

                # print(df[df['name'] == branch])
                logger.info(df[df['name'] == branch]['data'].iloc[0])
                logger.info(dick)
                # continue
                send_message_to_tg(tg_token, chat_id, f"Филиал, {df[df['name'] == branch]['name'].iloc[0]}")
                send_message_to_tg(tg_token, chat_id, f"Данные, {df[df['name'] == branch]['data'].iloc[0]}")
                facts = df[df['name'] == branch]['data'].iloc[0]

                first_val = int(data[0]) if data[0] is not None else 0
                second_val = int(data[1]) if data[1] is not None else 0
                third_val = int(data[2]) if data[2] is not None else 0

                try:
                    insert_data_in_db(started_time=datetime.datetime.now().strftime("%d.%m.%Y %H:%M:%S.%f"), store_id=id_, store_name=branch, full_name=full_name,
                                      executor_name=ip_address, status_='processing', error_reason='', error_saved_path='', execution_time='', ecp_path_=ecp_path, fact1=first_val, fact2=second_val, fact3=third_val, site1='', site2='', site3='')

                    status, saved_path, sites = start_single_branch(ecp_path, branch, facts, dick)
                    if sites == '':
                        sites = [''] * 3
                except Exception as poebotnya:

                    logger.warning(f'Fucking error occured1: {poebotnya}')
                    end_time = time.time()
                    insert_data_in_db(started_time=datetime.datetime.now().strftime("%d.%m.%Y %H:%M:%S.%f"), store_id=id_, store_name=branch, full_name=full_name,
                                      executor_name=ip_address, status_='slomalsya', error_reason=str(poebotnya)[:200], error_saved_path='', execution_time=str(end_time - start_time), ecp_path_=ecp_path, fact1=None, fact2=None, fact3=None, site1=None, site2=None, site3=None)
                    continue

                end_time = time.time()
                if status != 'success':
                    insert_data_in_db(started_time=datetime.datetime.now().strftime("%d.%m.%Y %H:%M:%S.%f"), store_id=id_, store_name=branch, full_name=full_name,
                                      executor_name=ip_address, status_='failed', error_reason=status, error_saved_path=saved_path, execution_time=str(end_time - start_time), ecp_path_=ecp_path, fact1=first_val, fact2=second_val, fact3=third_val, site1=sites[0], site2=sites[1], site3=sites[2])
                    # insert_data_in_db(started_time=start, store_name=str(branch), executor_name=ip_address, status_='failed', error_reason=status, error_saved_path=saved_path, execution_time=str(end_time - start_time), ecp_path_=ecp_path, fact1=facts[0], fact2=facts[1], fact3=facts[2], site1=sites[0], site2=sites[1], site3=sites[2])
                else:
                    insert_data_in_db(started_time=datetime.datetime.now().strftime("%d.%m.%Y %H:%M:%S.%f"), store_id=id_, store_name=branch, full_name=full_name,
                                      executor_name=ip_address, status_='success', error_reason=status, error_saved_path=saved_path, execution_time=str(end_time - start_time), ecp_path_=ecp_path, fact1=first_val, fact2=second_val, fact3=third_val, site1=sites[0], site2=sites[1], site3=sites[2])

                # send_message_to_tg(tg_token, chat_id, f'Finished, {ind}, {branch}')

            except Exception as ebanko:
                send_message_to_tg(tg_token, chat_id, f'Fucking error occured: {ebanko}')
                logger.warning(f'Fucking error occured: {ebanko}')
                end_time = time.time()
                saved_path = save_screenshot(df['name'].iloc[ind])
                insert_data_in_db(started_time=datetime.datetime.now().strftime("%d.%m.%Y %H:%M:%S.%f"), store_id=id_, store_name=branch, full_name=full_name,
                                  executor_name=ip_address, status_='slomalsya', error_reason=str(ebanko)[:200], error_saved_path=saved_path, execution_time=str(end_time - start_time), ecp_path_=ecp_path, fact1=None, fact2=None, fact3=None, site1=None, site2=None, site3=None)

    # print(ck)
    print(c)
    print(a)
    print(len(a))
    # exit()
    create_and_send_final_report()
    send_message_to_tg(tg_token, chat_id, f'Отработка заполнения стат отчёта 2Т Закончена!')
